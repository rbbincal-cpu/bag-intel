"""Daily Excel export: exports/bag-intel-YYYY-MM-DD.xlsx
Sheets: Summary, Hero Models, one per store."""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import ROOT

GOLD = "B08D3E"
HEAD_FILL = PatternFill("solid", fgColor="F5EFE3")
HEAD_FONT = Font(bold=True, color="6B5318")
TITLE_FONT = Font(bold=True, size=14)
PESO = '#,##0'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left")


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    with open(os.path.join(ROOT, "site", "data.json"), encoding="utf-8") as f:
        data = json.load(f)
    day = data["generated"]
    wb = Workbook()

    # ── Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Market summary — {day}"
    ws["A1"].font = TITLE_FONT
    cols = ["Store", "Items", "Inventory value", "Avg ticket", "Sold MTD",
            "Sold value MTD", "New MTD", "New value MTD", "Avg days to sell",
            "Aged 60d+ (count)", "Aged 60d+ (value)", "Aged 90d+ (count)",
            "Aged 90d+ (value)", "Markdowns MTD", "Avg discount %",
            "% Hermès", "% Chanel", "% LV", "% Other"]
    ws.append([])
    ws.append(cols)
    style_header(ws, 3, len(cols))
    for s in data["sites"]:
        bm = s["brand_mix"]
        ws.append([s["name"] + (" ★" if s["is_mine"] else ""), s["inventory_count"],
                   s["inventory_value"], s["avg_ticket"], s["sold_mtd_count"],
                   s["sold_mtd_value"], s["new_mtd_count"], s["new_mtd_value"],
                   s["avg_days_to_sell"], s["aging60_count"], s["aging60_value"],
                   s["aging90_count"], s["aging90_value"], s["markdowns_mtd"],
                   s["avg_discount_pct"], bm.get("Hermès", 0), bm.get("Chanel", 0),
                   bm.get("Louis Vuitton", 0), bm.get("Other", 0)])
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=8):
        for c in row:
            c.number_format = PESO
    autosize(ws, [22] + [14] * (len(cols) - 1))

    # ── Hero Models ────────────────────────────────────────────────
    ws = wb.create_sheet("Hero Models")
    ws["A1"] = f"Hero models — {day}"
    ws["A1"].font = TITLE_FONT
    cols = ["Model", "Store", "Live listings", "Sold MTD", "Avg sold price",
            "Avg days to sell", "Lowest ask (market)", "Highest ask (market)"]
    ws.append([])
    ws.append(cols)
    style_header(ws, 3, len(cols))
    for h in data["heroes"]:
        for key, v in h["per_site"].items():
            ws.append([h["name"], v["site_name"] + (" ★" if v["is_mine"] else ""),
                       len(v["listings"]), v["sold_mtd"], v["avg_sold_price"],
                       v["avg_days_to_sell"], h["lowest_ask"], h["highest_ask"]])
    for row in ws.iter_rows(min_row=4, min_col=5, max_col=8):
        for c in row:
            c.number_format = PESO
    autosize(ws, [26, 22, 14, 10, 14, 14, 16, 16])

    # ── Per-store sheets ───────────────────────────────────────────
    for s in data["sites"]:
        det = data["competitor_detail"][s["key"]]
        ws = wb.create_sheet(s["name"][:28])
        ws["A1"] = f'{s["name"]} — live listings ({day})'
        ws["A1"].font = TITLE_FONT
        cols = ["Title", "Brand", "Model", "Category", "Price", "Color",
                "Hardware", "Leather", "Size", "Days listed", "Status", "Hero", "URL"]
        ws.append([])
        ws.append(cols)
        style_header(ws, 3, len(cols))
        for p in det["listings"]:
            ws.append([p["title"], p["brand"], p["model"], p["category"],
                       p["price"], p["color"], p["hardware"], p["leather"],
                       p["size"], p["age_days"], p["status"], p["hero"], p["url"]])
        for row in ws.iter_rows(min_row=4, min_col=5, max_col=5):
            for c in row:
                c.number_format = PESO
        r0 = ws.max_row + 2
        ws.cell(row=r0, column=1, value="Recent sales detected").font = Font(bold=True)
        ws.append([])
        hdr = ["Title", "Brand", "Sold price", "Sold date", "Days to sell"]
        ws.append(hdr)
        style_header(ws, ws.max_row, len(hdr))
        for p in det["recent_sold"]:
            ws.append([p["title"], p["brand"], p["sold_price"], p["sold_date"],
                       p["days_to_sell"]])
        autosize(ws, [52, 14, 18, 12, 12, 14, 14, 14, 10, 11, 10, 20, 40])

    os.makedirs(os.path.join(ROOT, "exports"), exist_ok=True)
    out = os.path.join(ROOT, "exports", f"bag-intel-{day}.xlsx")
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
